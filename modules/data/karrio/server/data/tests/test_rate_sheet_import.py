"""
API tests for rate sheet import/export — flat single-sheet format.

Covers:
 1. test_parse_flat_xlsx_basic
 2. test_validate_flat_data_valid
 3. test_missing_service_rates_sheet_raises
 4. test_import_validation_errors_returned_before_write
 5. test_invalid_shipment_type_rejected
 6. test_blank_shipment_type_defaults_to_outbound
 7. test_duplicate_rate_rows_rejected
 8. test_missing_required_column_rejected
 9. test_build_rate_sheet_input_from_flat
10. test_compute_diff_all_new
11. test_dry_run_does_not_write
12. test_export_generates_flat_xlsx
13. test_export_reimport_roundtrip
14. test_import_csv_flat_format

Run with:
    karrio test karrio.server.data.tests.test_rate_sheet_import
  or:
    LOG_LEVEL=30 python -m unittest karrio.server.data.tests.test_rate_sheet_import -v
"""

import csv
import io
import unittest

import karrio.server.data.resources.rate_sheets as rs
import karrio.server.data.serializers.batch_rate_sheets as batch_rs
import openpyxl

# ─────────────────────────────────────────────────────────────────────────────
# Flat-format fixture builders
# ─────────────────────────────────────────────────────────────────────────────

# All columns in the exact flat-format order
FLAT_HEADERS = [
    "carrier_name",
    "service_code",
    "carrier_service_code",
    "service_name",
    "shipment_type",
    "origin_country",
    "zone_label",
    "country_codes",
    "min_weight",
    "max_weight",
    "weight_unit",
    "max_length",
    "max_width",
    "max_height",
    "dimension_unit",
    "currency",
    "base_rate",
    "cost",
    "transit_days",
    "transit_time",
    "plan_rate_start",
    "plan_cost_start",
    "plan_rate_advanced",
    "plan_cost_advanced",
    "plan_rate_pro",
    "plan_cost_pro",
    "plan_rate_enterprise",
    "plan_cost_enterprise",
    "tracked",
    "b2c",
    "b2b",
    "first_mile",
    "last_mile",
    "form_factor",
    "signature",
    "fuel_surcharge",
    "seasonal_surcharge",
    "customs_surcharge",
    "energy_surcharge",
    "road_toll",
    "security_surcharge",
    "notes",
]


def _flat_row(
    carrier_name="DHL Parcel DE",
    service_code="dhl_paket",
    carrier_service_code="V01PAK",
    service_name="DHL Paket",
    shipment_type="outbound",
    origin_country="DE",
    zone_label="Germany",
    country_codes="DE",
    min_weight=0,
    max_weight=2.0,
    weight_unit="KG",
    max_length=None,
    max_width=None,
    max_height=None,
    dimension_unit="CM",
    currency="EUR",
    base_rate=7.50,
    cost=4.80,
    transit_days=2,
    transit_time=None,
    **kwargs,
) -> list:
    """Return a list of values in FLAT_HEADERS order."""
    base = {
        "carrier_name": carrier_name,
        "service_code": service_code,
        "carrier_service_code": carrier_service_code,
        "service_name": service_name,
        "shipment_type": shipment_type,
        "origin_country": origin_country,
        "zone_label": zone_label,
        "country_codes": country_codes,
        "min_weight": min_weight,
        "max_weight": max_weight,
        "weight_unit": weight_unit,
        "max_length": max_length,
        "max_width": max_width,
        "max_height": max_height,
        "dimension_unit": dimension_unit,
        "currency": currency,
        "base_rate": base_rate,
        "cost": cost,
        "transit_days": transit_days,
        "transit_time": transit_time,
    }
    base.update(kwargs)
    return [base.get(h) for h in FLAT_HEADERS]


def _make_flat_xlsx(service_rate_rows=None) -> bytes:
    """Build a minimal valid flat-format xlsx and return bytes."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "service_rates"
    ws.append(FLAT_HEADERS)
    for row in service_rate_rows or [_flat_row()]:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_flat_csv(rows=None) -> bytes:
    """Build a flat-format CSV and return bytes."""
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(FLAT_HEADERS)
    for row in rows or [_flat_row()]:
        writer.writerow(row)
    return buf.getvalue().encode("utf-8")


def _make_file(data: bytes, name="test.xlsx"):
    """Wrap bytes in a file-like object."""
    f = io.BytesIO(data)
    f.name = name
    return f


# ─────────────────────────────────────────────────────────────────────────────
# Tests
# ─────────────────────────────────────────────────────────────────────────────


class TestRateSheetFlatFormat(unittest.TestCase):
    """Tests for the flat-format parser, validator, builder, diff, and export."""

    # ── 1 ────────────────────────────────────────────────────────────────────

    def test_parse_flat_xlsx_basic(self):
        """parse_flat_workbook should return a list of row-dicts."""
        data = _make_flat_xlsx()
        rows = rs.parse_flat_workbook(data)
        self.assertIsInstance(rows, list)
        self.assertEqual(len(rows), 1)
        row = rows[0]
        self.assertEqual(row["carrier_name"], "DHL Parcel DE")
        self.assertEqual(row["service_code"], "dhl_paket")
        self.assertEqual(row["zone_label"], "Germany")
        self.assertAlmostEqual(float(row["base_rate"]), 7.50)

    # ── 2 ────────────────────────────────────────────────────────────────────

    def test_validate_flat_data_valid(self):
        """Valid flat rows should produce no validation errors."""
        data = _make_flat_xlsx()
        rows = rs.parse_flat_workbook(data)
        errors = rs.validate_flat_data(rows)
        self.assertEqual(errors, [], f"Unexpected errors: {errors}")

    # ── 3 ────────────────────────────────────────────────────────────────────

    def test_missing_service_rates_sheet_raises(self):
        """A workbook without 'service_rates' sheet should raise ValueError."""
        wb = openpyxl.Workbook()
        buf = io.BytesIO()
        wb.save(buf)
        with self.assertRaises(ValueError):
            rs.parse_flat_workbook(buf.getvalue())

    # ── 4 ────────────────────────────────────────────────────────────────────

    def test_import_validation_errors_returned_before_write(self):
        """Missing required field (carrier_name) should surface as a validation error."""
        bad_row = _flat_row(carrier_name=None)  # carrier_name is None
        data = _make_flat_xlsx([bad_row])
        rows = rs.parse_flat_workbook(data)
        errors = rs.validate_flat_data(rows)
        self.assertTrue(
            any(e["field"] == "carrier_name" for e in errors),
            f"Expected carrier_name error, got: {errors}",
        )

    # ── 5 ────────────────────────────────────────────────────────────────────

    def test_invalid_shipment_type_rejected(self):
        """Invalid shipment_type value should appear in validation errors."""
        bad_row = _flat_row(shipment_type="air_mail")
        data = _make_flat_xlsx([bad_row])
        rows = rs.parse_flat_workbook(data)
        errors = rs.validate_flat_data(rows)
        self.assertTrue(
            any(e["field"] == "shipment_type" for e in errors),
            f"Expected shipment_type error, got: {errors}",
        )

    # ── 6 ────────────────────────────────────────────────────────────────────

    def test_blank_shipment_type_defaults_to_outbound(self):
        """
        A blank shipment_type is not in the invalid-values set so should pass.
        (The builder defaults it to 'outbound' at upsert time.)
        """
        row = _flat_row(shipment_type="")
        data = _make_flat_xlsx([row])
        rows = rs.parse_flat_workbook(data)
        errors = rs.validate_flat_data(rows)
        # A blank value should not trigger the shipment_type error
        self.assertFalse(
            any(e["field"] == "shipment_type" for e in errors),
            f"Did not expect shipment_type error, got: {errors}",
        )

    # ── 7 ────────────────────────────────────────────────────────────────────

    def test_duplicate_rate_rows_rejected(self):
        """Exact duplicate (carrier, service, zone, shipment_type, weights) should error."""
        row1 = _flat_row(base_rate=7.50)
        row2 = _flat_row(base_rate=8.00)  # same key, different rate
        data = _make_flat_xlsx([row1, row2])
        rows = rs.parse_flat_workbook(data)
        errors = rs.validate_flat_data(rows)
        self.assertTrue(
            any("Duplicate" in e["message"] for e in errors),
            f"Expected duplicate error, got: {errors}",
        )

    # ── 8 ────────────────────────────────────────────────────────────────────

    def test_missing_required_column_rejected(self):
        """Sheet missing a required column header should produce a column-level error."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "service_rates"
        # Write headers with 'currency' omitted
        incomplete = [h for h in FLAT_HEADERS if h != "currency"]
        ws.append(incomplete)
        ws.append([_flat_row()[FLAT_HEADERS.index(h)] for h in incomplete])
        buf = io.BytesIO()
        wb.save(buf)

        rows = rs.parse_flat_workbook(buf.getvalue())
        errors = rs.validate_flat_data(rows)
        self.assertTrue(
            any(e["field"] == "currency" for e in errors),
            f"Expected currency column error, got: {errors}",
        )

    # ── 9 ────────────────────────────────────────────────────────────────────

    def test_build_rate_sheet_input_from_flat(self):
        """build_rate_sheet_input_from_flat should group rows by carrier correctly."""
        rows_data = [
            _flat_row(zone_label="Germany", country_codes="DE", min_weight=0, max_weight=2, base_rate=7.50),
            _flat_row(zone_label="France", country_codes="FR", min_weight=0, max_weight=2, base_rate=9.00),
            _flat_row(
                service_code="dhl_paket_xl",
                carrier_service_code="V53PAK",
                service_name="DHL Paket XL",
                zone_label="Germany",
                country_codes="DE",
                min_weight=2,
                max_weight=5,
                base_rate=10.50,
            ),
        ]
        data = _make_flat_xlsx(rows_data)
        rows = rs.parse_flat_workbook(data)
        carrier_inputs = rs.build_rate_sheet_input_from_flat(rows)

        self.assertIn("DHL Parcel DE", carrier_inputs)
        ci = carrier_inputs["DHL Parcel DE"]
        self.assertEqual(len(ci["zones"]), 2, f"Expected 2 zones, got {ci['zones']}")
        self.assertEqual(len(ci["services"]), 2, f"Expected 2 services, got {ci['services']}")
        self.assertEqual(len(ci["service_rates"]), 3)

        # Check payload conversion
        payload = rs.flat_carrier_to_upsert_payload(ci)
        self.assertEqual(payload["rate_sheet"]["carrier_name"], "DHL Parcel DE")
        self.assertEqual(len(payload["service_rates"]), 3)
        self.assertAlmostEqual(payload["service_rates"][0]["rate"], 7.50)

    # ── 10 ───────────────────────────────────────────────────────────────────

    def test_compute_diff_all_new(self):
        """compute_diff with no existing sheet should mark all rows as added."""
        data = _make_flat_xlsx()
        rows = rs.parse_flat_workbook(data)
        errors = rs.validate_flat_data(rows)
        self.assertEqual(errors, [])

        carrier_inputs = rs.build_rate_sheet_input_from_flat(rows)
        carrier_name = next(iter(carrier_inputs))
        payload = rs.flat_carrier_to_upsert_payload(carrier_inputs[carrier_name])

        diff = rs.compute_diff(None, payload)
        self.assertEqual(diff["summary"]["added"], len(payload["service_rates"]))
        self.assertEqual(diff["summary"]["updated"], 0)
        self.assertTrue(all(r["change"] == "added" for r in diff["rows"]))
        # New-format fields present
        self.assertIn("service_code", diff["rows"][0])
        self.assertIn("zone_label", diff["rows"][0])

    # ── 11 ───────────────────────────────────────────────────────────────────

    def test_dry_run_does_not_write(self):
        """dry_run=True with valid flat file returns diff without any DB writes."""

        class FakeContext:
            test_mode = True
            user = None
            org = None

        data = _make_flat_xlsx()
        f = _make_file(data, "test.xlsx")

        from unittest.mock import MagicMock, patch

        mock_qs = MagicMock()
        mock_qs.filter.return_value.first.return_value = None

        with patch("karrio.server.providers.models.RateSheet.access_by", return_value=mock_qs):
            result = batch_rs.process_rate_sheet_import(
                data_file=f,
                context=FakeContext(),
                dry_run=True,
            )

        self.assertTrue(result.get("dry_run"))
        self.assertIn("diff", result)
        self.assertEqual(result.get("errors", []), [])
        # Diff summary should show at least 1 added row
        self.assertGreater(result["diff"]["summary"]["added"], 0)

    # ── 12 ───────────────────────────────────────────────────────────────────

    def test_export_generates_flat_xlsx(self):
        """export_rate_sheet_xlsx should produce a flat workbook with one sheet."""

        class MockService:
            id = "dhl_paket"
            service_name = "DHL Paket"
            service_code = "dhl_paket"
            carrier_service_code = "V01PAK"
            currency = "EUR"
            features = {"b2c": True, "b2b": False}
            domicile = True
            international = False
            tracked = True
            max_weight = 31.5
            min_weight = None
            max_length = 120.0
            max_width = 60.0
            max_height = 60.0
            dimension_unit = "CM"
            weight_unit = "KG"
            zone_ids = ["Germany"]
            surcharge_ids = []

        class MockServices:
            def all(self):
                return [MockService()]

        class MockSheet:
            id = "rsht_test_1"
            name = "DHL Parcel DE"
            carrier_name = "dhl_parcel_de"
            slug = "dhl_parcel_de"
            origin_countries = ["DE"]
            currency = "EUR"
            zones = [{"id": "Germany", "label": "Germany", "country_codes": ["DE"]}]
            surcharges = []
            service_rates = [
                {
                    "service_id": "dhl_paket",
                    "zone_id": "Germany",
                    "shipment_type": "outbound",
                    "min_weight": 0,
                    "max_weight": 2.0,
                    "rate": 7.50,
                    "cost": 4.80,
                    "transit_days": "2",
                }
            ]
            services = MockServices()

        xlsx_bytes = rs.export_rate_sheet_xlsx(MockSheet())
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))

        # New flat format: single sheet named 'service_rates'
        self.assertEqual(wb.sheetnames, ["service_rates"])
        ws = wb["service_rates"]

        # Check header row
        headers = [cell.value for cell in ws[1]]
        self.assertIn("carrier_name", headers)
        self.assertIn("service_code", headers)
        self.assertIn("zone_label", headers)
        self.assertIn("base_rate", headers)

        # Check data row exists
        data_rows = list(ws.iter_rows(min_row=2, values_only=True))
        self.assertEqual(len(data_rows), 1)
        row_dict = dict(zip(headers, data_rows[0], strict=False))
        self.assertEqual(row_dict["carrier_name"], "dhl_parcel_de")
        self.assertEqual(row_dict["service_code"], "dhl_paket")
        self.assertEqual(row_dict["zone_label"], "Germany")
        self.assertAlmostEqual(float(row_dict["base_rate"]), 7.50)

    # ── 13 ───────────────────────────────────────────────────────────────────

    def test_export_reimport_roundtrip(self):
        """Export a flat sheet then re-parse it — service_rates should be stable."""

        class MockService:
            id = "dhl_paket"
            service_name = "DHL Paket"
            service_code = "dhl_paket"
            carrier_service_code = "V01PAK"
            currency = "EUR"
            features = {}
            domicile = True
            international = False
            tracked = True
            max_weight = 31.5
            min_weight = None
            max_length = None
            max_width = None
            max_height = None
            dimension_unit = "CM"
            weight_unit = "KG"
            zone_ids = ["Germany"]
            surcharge_ids = []

        class MockServices:
            def all(self):
                return [MockService()]

        class MockSheet:
            id = "rsht_test_2"
            name = "DHL Parcel DE"
            carrier_name = "dhl_parcel_de"
            slug = "dhl_parcel_de"
            origin_countries = ["DE"]
            currency = "EUR"
            zones = [{"id": "Germany", "label": "Germany", "country_codes": ["DE"]}]
            surcharges = []
            service_rates = [
                {
                    "service_id": "dhl_paket",
                    "zone_id": "Germany",
                    "shipment_type": "outbound",
                    "min_weight": 0,
                    "max_weight": 2.0,
                    "rate": 7.50,
                    "cost": 4.80,
                },
                {
                    "service_id": "dhl_paket",
                    "zone_id": "Germany",
                    "shipment_type": "returns",
                    "min_weight": 0,
                    "max_weight": 2.0,
                    "rate": 8.00,
                    "cost": 5.00,
                },
            ]
            services = MockServices()

        # Export → flat xlsx
        xlsx_bytes = rs.export_rate_sheet_xlsx(MockSheet())

        # Verify it's recognised as flat format
        wb_fmt = rs.detect_workbook_format(xlsx_bytes)
        self.assertEqual(wb_fmt, "flat", f"Expected flat format, got {wb_fmt!r}")

        # Re-import
        rows = rs.parse_flat_workbook(xlsx_bytes)
        self.assertEqual(len(rows), 2)

        errors = rs.validate_flat_data(rows)
        self.assertEqual(errors, [], f"Roundtrip validation errors: {errors}")

        carrier_inputs = rs.build_rate_sheet_input_from_flat(rows)
        self.assertTrue(len(carrier_inputs) >= 1)

        carrier_name = next(iter(carrier_inputs))
        ci = carrier_inputs[carrier_name]
        self.assertEqual(len(ci["service_rates"]), 2)

        # Verify rates preserved
        {sr["zone_label"]: sr for sr in ci["service_rates"]}
        # Both rows have same zone; distinguish by rate value
        rate_vals = sorted([sr.get("rate") for sr in ci["service_rates"]])
        self.assertAlmostEqual(rate_vals[0], 7.50)
        self.assertAlmostEqual(rate_vals[1], 8.00)

    # ── 14 ───────────────────────────────────────────────────────────────────

    def test_import_csv_flat_format(self):
        """Flat CSV (with carrier_name column) should parse and validate without errors."""
        csv_data = _make_flat_csv()
        rows = rs._csv_rows(csv_data)

        self.assertIn("carrier_name", rows[0])
        self.assertEqual(len(rows), 1)

        errors = rs.validate_flat_data(rows)
        self.assertEqual(errors, [], f"Unexpected CSV errors: {errors}")

        carrier_inputs = rs.build_rate_sheet_input_from_flat(rows)
        self.assertIn("DHL Parcel DE", carrier_inputs)
        ci = carrier_inputs["DHL Parcel DE"]
        self.assertEqual(len(ci["service_rates"]), 1)
        self.assertAlmostEqual(float(ci["service_rates"][0]["rate"] or 0), 7.50)


if __name__ == "__main__":
    unittest.main()
