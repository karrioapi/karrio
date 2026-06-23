"""
Rate sheet import/export resource.

Supports two workbook formats:
  - NEW (flat): single `service_rates` sheet — every row is a fully-denormalised
    (carrier × service × zone × weight_range) record.  Detected when the workbook
    has a `service_rates` sheet but NO `zones` sheet.
  - LEGACY (multi-sheet): 5-6 sheet template (rate_sheet, zones, services,
    surcharges, service_rates).  Kept as a fallback.

CSV imports are considered flat-format when the first row contains `carrier_name`.
"""

import csv
import io

import karrio.lib as lib
import openpyxl

# ─────────────────────────────────────────────────────────────────────────────
# FLAT FORMAT COLUMN DEFINITIONS
# ─────────────────────────────────────────────────────────────────────────────

REQUIRED_COLUMNS = [
    "carrier_name",
    "service_code",
    "service_name",
    "shipment_type",
    "origin_country",
    "zone_label",
    "min_weight",
    "max_weight",
    "currency",
]

OPTIONAL_COLUMNS = [
    "carrier_service_code",
    "country_codes",
    "weight_unit",
    "max_length",
    "max_width",
    "max_height",
    "dimension_unit",
    "base_rate",
    "cost",
    "transit_days",
    "transit_time",
    "plan_rate_start",
    "plan_cost_start",
    "plan_rate_advanced",
    "plan_cost_advanced",
    "plan_rate_pro",
    "plan_cost_pro",
    "plan_rate_enterprise",
    "plan_cost_enterprise",
    "tracked",
    "b2c",
    "b2b",
    "first_mile",
    "last_mile",
    "form_factor",
    "signature",
    "fuel_surcharge",
    "seasonal_surcharge",
    "customs_surcharge",
    "energy_surcharge",
    "road_toll",
    "security_surcharge",
    "notes",
]

# ─────────────────────────────────────────────────────────────────────────────
# LEGACY FORMAT COLUMN DEFINITIONS (kept for backward-compat)
# ─────────────────────────────────────────────────────────────────────────────

DEFAULT_HEADERS = {
    "service_id": "service_id",
    "zone_id": "zone_id",
    "shipment_type": "shipment_type",
    "min_weight": "min_weight",
    "max_weight": "max_weight",
    "base_rate": "base_rate",
    "cost": "cost",
}

VALID_SHIPMENT_TYPES = {"outbound", "returns", "pickup"}

RATE_SHEET_COLS = ["name", "carrier_name", "slug", "currency", "origin_countries"]
ZONE_COLS = ["zone_id", "zone_label", "country_codes", "postal_codes", "cities", "transit_days", "transit_time"]
SERVICE_COLS = [
    "service_id",
    "service_name",
    "service_code",
    "carrier_service_code",
    "currency",
    "shipment_type",
    "domicile",
    "international",
    "tracked",
    "b2c",
    "b2b",
    "first_mile",
    "last_mile",
    "form_factor",
    "address_validation",
    "notification",
    "ddp_available",
    "ddu_available",
    "surcharge_ids",
    "max_weight",
    "min_weight",
    "max_length",
    "max_width",
    "max_height",
    "dimension_unit",
    "weight_unit",
]
SURCHARGE_COLS = ["surcharge_id", "surcharge_name", "amount", "surcharge_type", "cost", "active"]
SERVICE_RATE_COLS = [
    "service_id",
    "zone_id",
    "shipment_type",
    "min_weight",
    "max_weight",
    "base_rate",
    "cost",
    "transit_days",
    "plan_rate_start",
    "plan_cost_start",
    "plan_rate_advanced",
    "plan_cost_advanced",
    "plan_rate_pro",
    "plan_cost_pro",
    "plan_rate_enterprise",
    "plan_cost_enterprise",
    "surcharge_ids",
]

REQUIRED_RATE_SHEET_COLS = ["name", "carrier_name", "slug"]
REQUIRED_ZONE_COLS = ["zone_id", "zone_label"]
REQUIRED_SERVICE_COLS = ["service_id", "service_name", "service_code"]
REQUIRED_SURCHARGE_COLS = ["surcharge_id", "surcharge_name", "amount", "surcharge_type"]
REQUIRED_SERVICE_RATE_COLS = ["service_id", "zone_id", "base_rate"]


class ParseError(Exception):
    """Structured validation error list."""

    def __init__(self, errors: list):
        self.errors = errors
        super().__init__(str(errors))


# ─────────────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────────────


def _rows(ws) -> list:
    """Extract rows from an openpyxl worksheet as list of dicts."""
    headers = [cell.value for cell in ws[1]]
    result = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        result.append({headers[i]: row[i] for i in range(len(headers)) if i < len(headers)})
    return result


def _csv_rows(content: bytes) -> list:
    """Parse CSV bytes to list of dicts."""
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    return [row for row in reader if any(v.strip() for v in row.values())]


def _bool(val) -> bool | None:
    if val is None or val == "":
        return None
    if isinstance(val, bool):
        return val
    return str(val).lower() in ("true", "1", "yes")


def _float(val) -> float | None:
    if val is None or val == "":
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def _int(val) -> int | None:
    if val is None or val == "":
        return None
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return None


def _str(val) -> str | None:
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def _split_ids(val) -> list:
    if not val:
        return []
    return [s.strip() for s in str(val).split(",") if s.strip()]


def _parse_countries(val) -> list:
    if not val:
        return []
    return [s.strip().upper() for s in str(val).split(",") if s.strip()]


def _build_rate_meta(row: dict) -> dict:
    """Build rate metadata dict from a flat row (surcharges, plans, extras)."""
    meta = {}
    surcharge_fields = [
        "fuel_surcharge",
        "seasonal_surcharge",
        "customs_surcharge",
        "energy_surcharge",
        "road_toll",
        "security_surcharge",
    ]
    plan_fields = [
        "plan_rate_start",
        "plan_cost_start",
        "plan_rate_advanced",
        "plan_cost_advanced",
        "plan_rate_pro",
        "plan_cost_pro",
        "plan_rate_enterprise",
        "plan_cost_enterprise",
    ]
    for field in surcharge_fields + plan_fields:
        val = _float(row.get(field))
        if val is not None:
            meta[field] = val
    sig = row.get("signature")
    if sig is not None and sig != "":
        meta["signature"] = _bool(sig)
    tt = _str(row.get("transit_time"))
    if tt:
        meta["transit_time"] = tt
    notes = _str(row.get("notes"))
    if notes:
        meta["notes"] = notes
    return meta


# ─────────────────────────────────────────────────────────────────────────────
# FORMAT DETECTION
# ─────────────────────────────────────────────────────────────────────────────


def detect_format(data: bytes, filename: str) -> str:
    """Return 'xlsx', 'csv', or raise ValueError. (File-type detection.)"""
    name = (filename or "").lower()
    if name.endswith(".csv"):
        return "csv"
    if name.endswith(".xls"):
        return "xls"
    if data[:2] == b"PK":
        return "xlsx"
    try:
        text = data[:512].decode("utf-8-sig")
        csv.Sniffer().sniff(text)
        return "csv"
    except Exception:  # noqa: S110 — probe step, fall through to extension-based detection
        pass
    if name.endswith(".xlsx"):
        return "xlsx"
    return "xlsx"  # default


def detect_workbook_format(data: bytes) -> str:
    """
    Inspect an xlsx workbook's sheet names and return the format type.

    Returns:
        'flat'         — new single-sheet format (service_rates, no zones sheet)
        'multi_sheet'  — legacy 5-6 sheet format (service_rates + zones)
        'legacy'       — unrecognised / missing service_rates sheet
    """
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    names = set(wb.sheetnames)
    wb.close()
    if "service_rates" in names and "zones" not in names:
        return "flat"
    elif "service_rates" in names and "zones" in names:
        return "multi_sheet"
    else:
        return "legacy"


# ─────────────────────────────────────────────────────────────────────────────
# FLAT FORMAT — PARSER
# ─────────────────────────────────────────────────────────────────────────────


def parse_flat_workbook(data: bytes) -> list:
    """
    Parse the new single-sheet flat format.

    Expects a workbook with a single ``service_rates`` sheet.
    Returns a list of row-dicts (one per data row, blank rows skipped).
    """
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    if "service_rates" not in wb.sheetnames:
        raise ValueError(
            "Workbook does not contain a 'service_rates' sheet. Please use the current flat-format template."
        )
    ws = wb["service_rates"]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(v is not None for v in row):
            continue
        rows.append(dict(zip(headers, row, strict=False)))
    wb.close()
    return rows


# ─────────────────────────────────────────────────────────────────────────────
# FLAT FORMAT — VALIDATOR
# ─────────────────────────────────────────────────────────────────────────────


def validate_flat_data(rows: list) -> list:
    """
    Validate flat format rows.

    Returns a list of error dicts (empty list = valid).
    """
    errors = []
    if not rows:
        errors.append(
            {
                "sheet": "service_rates",
                "row": 0,
                "field": "",
                "message": "service_rates sheet is empty",
            }
        )
        return errors

    # Check required columns are present at all
    present = set(rows[0].keys())
    for col in REQUIRED_COLUMNS:
        if col not in present:
            errors.append(
                {
                    "sheet": "service_rates",
                    "row": 1,
                    "field": col,
                    "message": f"Required column '{col}' is missing from the sheet",
                }
            )
    if errors:
        return errors  # column errors make row-level checks meaningless

    # Fields that are numeric — 0 is a valid value; check for None only
    _NUMERIC_FIELDS = {"min_weight", "max_weight"}
    # Fields where blank string is acceptable (system uses a default)
    _OPTIONAL_BLANK = {"shipment_type"}

    seen_keys: dict = {}
    for i, row in enumerate(rows, start=2):
        for col in REQUIRED_COLUMNS:
            if col in _OPTIONAL_BLANK:
                continue  # blank → use builder default (e.g. "outbound")
            val = row.get(col)
            if col in _NUMERIC_FIELDS:
                missing = val is None
            else:
                missing = (
                    val is None
                    or (isinstance(val, str) and not val.strip())
                    or (not isinstance(val, (int, float, bool)) and not val)
                )
            if missing:
                errors.append(
                    {
                        "sheet": "service_rates",
                        "row": i,
                        "field": col,
                        "message": f"Required field '{col}' is missing",
                    }
                )

        # Validate shipment_type
        st = _str(row.get("shipment_type"))
        if st and st not in VALID_SHIPMENT_TYPES:
            errors.append(
                {
                    "sheet": "service_rates",
                    "row": i,
                    "field": "shipment_type",
                    "message": (f"Invalid shipment_type '{st}' — must be outbound, returns, or pickup"),
                }
            )

        # Validate weight range
        try:
            mn = float(row.get("min_weight") or 0)
            mx = float(row.get("max_weight") or 0)
            if mn > 0 and mx > 0 and mn >= mx:
                errors.append(
                    {
                        "sheet": "service_rates",
                        "row": i,
                        "field": "min_weight",
                        "message": "min_weight must be less than max_weight",
                    }
                )
        except (TypeError, ValueError):
            errors.append(
                {
                    "sheet": "service_rates",
                    "row": i,
                    "field": "min_weight",
                    "message": "min_weight/max_weight must be numeric",
                }
            )

        # Duplicate key check
        effective_st = st or "outbound"
        rate_key = (
            _str(row.get("carrier_name")),
            _str(row.get("service_code")),
            _str(row.get("zone_label")),
            effective_st,
            float(row.get("min_weight") or 0),
            float(row.get("max_weight") or 0),
        )
        if rate_key in seen_keys:
            prev = seen_keys[rate_key]
            errors.append(
                {
                    "sheet": "service_rates",
                    "row": i,
                    "field": "service_code",
                    "message": (
                        f"Duplicate rate row for {row.get('service_code')}/"
                        f"{row.get('zone_label')}/{row.get('min_weight')}-"
                        f"{row.get('max_weight')} kg (rows {prev}, {i})"
                    ),
                }
            )
        else:
            seen_keys[rate_key] = i

    return errors


# ─────────────────────────────────────────────────────────────────────────────
# FLAT FORMAT — BUILDER
# ─────────────────────────────────────────────────────────────────────────────


def build_rate_sheet_input_from_flat(rows: list) -> dict:
    """
    Build structured input from flat rows, grouped by carrier_name.

    Returns:
        {carrier_name: {"carrier_name", "origin_countries", "zones", "services", "service_rates"}}
    """
    from collections import defaultdict

    grouped = defaultdict(list)
    for row in rows:
        cn = _str(row.get("carrier_name"))
        if cn:
            grouped[cn].append(row)

    result = {}
    for carrier, carrier_rows in grouped.items():
        # Unique zones keyed by (zone_label, country_codes)
        zones: dict = {}
        for r in carrier_rows:
            zone_key = (_str(r.get("zone_label")), _str(r.get("country_codes")) or "")
            if zone_key not in zones:
                zones[zone_key] = {
                    "label": _str(r.get("zone_label")),
                    "country_codes": _parse_countries(r.get("country_codes")),
                    "transit_days": _str(r.get("transit_days")),
                }

        # Unique services keyed by service_code
        services: dict = {}
        for r in carrier_rows:
            sc = _str(r.get("service_code"))
            if sc and sc not in services:
                services[sc] = {
                    "service_name": _str(r.get("service_name")),
                    "service_code": sc,
                    "carrier_service_code": _str(r.get("carrier_service_code")),
                    "currency": _str(r.get("currency")) or "EUR",
                    "shipment_type": _str(r.get("shipment_type")) or "outbound",
                    "tracked": _bool(r.get("tracked")),
                    "b2c": _bool(r.get("b2c")),
                    "b2b": _bool(r.get("b2b")),
                    "first_mile": _str(r.get("first_mile")),
                    "last_mile": _str(r.get("last_mile")),
                    "form_factor": _str(r.get("form_factor")),
                    "max_weight": _float(r.get("max_weight")),
                    "max_length": _float(r.get("max_length")),
                    "max_width": _float(r.get("max_width")),
                    "max_height": _float(r.get("max_height")),
                    "weight_unit": _str(r.get("weight_unit")) or "KG",
                    "dimension_unit": _str(r.get("dimension_unit")) or "CM",
                }

        # Per-row service_rates
        service_rates = []
        for r in carrier_rows:
            service_rates.append(
                {
                    "service_code": _str(r.get("service_code")),
                    "zone_label": _str(r.get("zone_label")),
                    "min_weight": _float(r.get("min_weight")),
                    "max_weight": _float(r.get("max_weight")),
                    "rate": _float(r.get("base_rate")),
                    "cost": _float(r.get("cost")),
                    "transit_days": _int(r.get("transit_days")),
                    "meta": _build_rate_meta(r),
                }
            )

        result[carrier] = {
            "carrier_name": carrier,
            "origin_countries": list(
                {_str(r.get("origin_country")) for r in carrier_rows if _str(r.get("origin_country"))}
            ),
            "zones": list(zones.values()),
            "services": list(services.values()),
            "service_rates": service_rates,
        }

    return result


def flat_carrier_to_upsert_payload(carrier_input: dict) -> dict:
    """
    Convert flat carrier input (from build_rate_sheet_input_from_flat) to the
    standard upsert payload format accepted by upsert_rate_sheet().

    Convention: service_code is used as the service "id"; zone_label is used as
    the zone "id".  This keeps IDs human-readable and stable across re-imports.
    """
    carrier_name = carrier_input["carrier_name"]

    zones = [
        {
            "id": z["label"],
            "label": z["label"],
            "country_codes": z.get("country_codes", []),
            "postal_codes": [],
            "cities": [],
            "transit_days": _str(z.get("transit_days")),
            "transit_time": None,
        }
        for z in carrier_input["zones"]
    ]

    services = [
        {
            "id": s["service_code"],
            "service_name": s.get("service_name", ""),
            "service_code": s.get("service_code", ""),
            "carrier_service_code": s.get("carrier_service_code"),
            "currency": s.get("currency", "EUR"),
            "shipment_type": s.get("shipment_type", "outbound"),
            "domicile": None,
            "international": None,
            "tracked": s.get("tracked"),
            "b2c": s.get("b2c"),
            "b2b": s.get("b2b"),
            "first_mile": s.get("first_mile"),
            "last_mile": s.get("last_mile"),
            "form_factor": s.get("form_factor"),
            "max_weight": s.get("max_weight"),
            "min_weight": None,
            "max_length": s.get("max_length"),
            "max_width": s.get("max_width"),
            "max_height": s.get("max_height"),
            "weight_unit": s.get("weight_unit", "KG"),
            "dimension_unit": s.get("dimension_unit", "CM"),
            "zone_ids": [],
            "surcharge_ids": [],
        }
        for s in carrier_input["services"]
    ]

    # Populate zone_ids per service from service_rates
    svc_zone_map: dict = {}
    for sr in carrier_input["service_rates"]:
        sc = sr.get("service_code")
        zl = sr.get("zone_label")
        if sc and zl:
            svc_zone_map.setdefault(sc, set()).add(zl)
    for svc in services:
        sc = svc["service_code"]
        if sc in svc_zone_map:
            svc["zone_ids"] = list(svc_zone_map[sc])

    service_rates = []
    for sr in carrier_input["service_rates"]:
        meta = sr.get("meta") or {}
        service_rates.append(
            {
                "service_id": sr["service_code"],
                "zone_id": sr["zone_label"],
                "shipment_type": "outbound",  # default; per-row st carried in meta if needed
                "min_weight": sr.get("min_weight") or 0.0,
                "max_weight": sr.get("max_weight") or 0.0,
                "rate": sr.get("rate") or 0.0,
                "cost": sr.get("cost"),
                "transit_days": _str(sr.get("transit_days")),
                "plan_rate_start": meta.get("plan_rate_start"),
                "plan_cost_start": meta.get("plan_cost_start"),
                "plan_rate_advanced": meta.get("plan_rate_advanced"),
                "plan_cost_advanced": meta.get("plan_cost_advanced"),
                "plan_rate_pro": meta.get("plan_rate_pro"),
                "plan_cost_pro": meta.get("plan_cost_pro"),
                "plan_rate_enterprise": meta.get("plan_rate_enterprise"),
                "plan_cost_enterprise": meta.get("plan_cost_enterprise"),
                "surcharge_ids": [],
            }
        )

    currency = carrier_input["services"][0].get("currency", "EUR") if carrier_input["services"] else "EUR"
    slug = carrier_name.lower().replace(" ", "_").replace("-", "_")

    return {
        "rate_sheet": {
            "name": carrier_name,
            "carrier_name": carrier_name,
            "slug": slug,
            "currency": currency,
            "origin_countries": carrier_input.get("origin_countries", []),
        },
        "zones": zones,
        "services": services,
        "surcharges": [],
        "service_rates": service_rates,
    }


# ─────────────────────────────────────────────────────────────────────────────
# LEGACY FORMAT — VALIDATION
# ─────────────────────────────────────────────────────────────────────────────


def _check_required_cols(rows: list, required: list, sheet: str, errors: list):
    if not rows:
        return
    present = set(rows[0].keys())
    for col in required:
        if col not in present:
            errors.append(
                {
                    "sheet": sheet,
                    "row": 1,
                    "field": col,
                    "message": f"Missing required column '{col}' in sheet '{sheet}'",
                }
            )


def validate_workbook(parsed: dict) -> list:
    """Validate legacy multi-sheet workbook. Returns list of error dicts."""
    errors = []
    rs = parsed["rate_sheet"]
    zones = parsed["zones"]
    services = parsed["services"]
    surcharges = parsed["surcharges"]
    service_rates = parsed["service_rates"]

    _check_required_cols(rs if rs else [{}], REQUIRED_RATE_SHEET_COLS, "rate_sheet", errors)
    _check_required_cols(zones, REQUIRED_ZONE_COLS, "zones", errors)
    _check_required_cols(services, REQUIRED_SERVICE_COLS, "services", errors)
    if surcharges:
        _check_required_cols(surcharges, REQUIRED_SURCHARGE_COLS, "surcharges", errors)
    _check_required_cols(service_rates, REQUIRED_SERVICE_RATE_COLS, "service_rates", errors)

    if errors:
        return errors

    if not service_rates:
        errors.append(
            {
                "sheet": "service_rates",
                "row": 0,
                "field": "",
                "message": "service_rates sheet is empty",
            }
        )

    zone_ids = {r.get("zone_id") for r in zones}
    service_ids = {r.get("service_id") for r in services}
    surcharge_ids = {r.get("surcharge_id") for r in surcharges}

    seen_rate_keys: dict = {}

    for i, row in enumerate(service_rates, 2):
        zone_id = _str(row.get("zone_id"))
        service_id = _str(row.get("service_id"))
        shipment_type = _str(row.get("shipment_type")) or "outbound"
        min_w = _float(row.get("min_weight")) or 0.0
        max_w = _float(row.get("max_weight")) or 0.0

        if zone_id and zone_id not in zone_ids:
            errors.append(
                {
                    "sheet": "service_rates",
                    "row": i,
                    "field": "zone_id",
                    "message": f"zone_id '{zone_id}' not found in zones sheet (row {i})",
                }
            )

        if service_id and service_id not in service_ids:
            errors.append(
                {
                    "sheet": "service_rates",
                    "row": i,
                    "field": "service_id",
                    "message": f"service_id '{service_id}' not found in services sheet (row {i})",
                }
            )

        if shipment_type not in VALID_SHIPMENT_TYPES:
            errors.append(
                {
                    "sheet": "service_rates",
                    "row": i,
                    "field": "shipment_type",
                    "message": (
                        f"Invalid shipment_type '{shipment_type}' — must be outbound, returns, or pickup (row {i})"
                    ),
                }
            )

        if not (min_w == 0.0 and max_w == 0.0) and min_w >= max_w:
            errors.append(
                {
                    "sheet": "service_rates",
                    "row": i,
                    "field": "min_weight",
                    "message": f"min_weight must be less than max_weight (row {i})",
                }
            )

        for sid in _split_ids(row.get("surcharge_ids")):
            if surcharges and sid not in surcharge_ids:
                errors.append(
                    {
                        "sheet": "service_rates",
                        "row": i,
                        "field": "surcharge_ids",
                        "message": f"surcharge_id '{sid}' not found in surcharges sheet (row {i})",
                    }
                )

        rate_key = (service_id, zone_id, shipment_type, min_w, max_w)
        if rate_key in seen_rate_keys:
            prev = seen_rate_keys[rate_key]
            errors.append(
                {
                    "sheet": "service_rates",
                    "row": i,
                    "field": "service_id",
                    "message": (f"Duplicate rate row for {service_id}/{zone_id}/{min_w}-{max_w} kg (rows {prev}, {i})"),
                }
            )
        else:
            seen_rate_keys[rate_key] = i

    return errors


# ─────────────────────────────────────────────────────────────────────────────
# LEGACY FORMAT — PARSERS
# ─────────────────────────────────────────────────────────────────────────────


def parse_xlsx(data: bytes) -> dict:
    """Parse legacy 5-6-sheet Excel workbook. Returns dict of sheet_name → list[dict]."""
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)

    if "service_rates" not in wb.sheetnames:
        raise ValueError(
            "This workbook does not appear to be a karrio rate sheet template "
            "(missing 'service_rates' sheet). Use the template from templates/ folder."
        )

    parsed = {
        "rate_sheet": _rows(wb["rate_sheet"]) if "rate_sheet" in wb.sheetnames else [],
        "zones": _rows(wb["zones"]) if "zones" in wb.sheetnames else [],
        "services": _rows(wb["services"]) if "services" in wb.sheetnames else [],
        "surcharges": _rows(wb["surcharges"]) if "surcharges" in wb.sheetnames else [],
        "service_rates": _rows(wb["service_rates"]),
    }
    return parsed


def parse_csv(data: bytes, context=None, rate_sheet_id: str = None) -> dict:
    """
    Parse a flat CSV.

    If the CSV has a ``carrier_name`` column it is treated as a flat-format CSV;
    otherwise it is treated as a legacy service_rates-only CSV.

    Returns a partial parsed dict compatible with validate_workbook / the flat
    validation path.
    """
    rows = _csv_rows(data)
    return {
        "rate_sheet": [],
        "zones": [],
        "services": [],
        "surcharges": [],
        "service_rates": rows,
        "_csv_only": True,
        "_rate_sheet_ref": rate_sheet_id,
    }


# ─────────────────────────────────────────────────────────────────────────────
# LEGACY FORMAT — PAYLOAD BUILDER
# ─────────────────────────────────────────────────────────────────────────────


def build_upsert_payload(parsed: dict) -> dict:
    """
    Convert legacy parsed workbook data to the upsert payload dict.
    """
    rs_rows = parsed.get("rate_sheet", [])
    rs_row = rs_rows[0] if rs_rows else {}

    zones = [
        {
            "id": _str(r.get("zone_id")),
            "label": _str(r.get("zone_label")),
            "country_codes": _parse_countries(r.get("country_codes")),
            "postal_codes": _split_ids(r.get("postal_codes")),
            "cities": _split_ids(r.get("cities")),
            "transit_days": _str(r.get("transit_days")),
            "transit_time": _float(r.get("transit_time")),
        }
        for r in parsed.get("zones", [])
        if _str(r.get("zone_id"))
    ]

    services = [
        {
            "id": _str(r.get("service_id")),
            "service_name": _str(r.get("service_name")),
            "service_code": (_str(r.get("service_code")) or "").lower().replace(" ", "_"),
            "carrier_service_code": _str(r.get("carrier_service_code")),
            "currency": _str(r.get("currency")),
            "shipment_type": _str(r.get("shipment_type")) or "outbound",
            "domicile": _bool(r.get("domicile")),
            "international": _bool(r.get("international")),
            "tracked": _bool(r.get("tracked")),
            "b2c": _bool(r.get("b2c")),
            "b2b": _bool(r.get("b2b")),
            "first_mile": _str(r.get("first_mile")),
            "last_mile": _str(r.get("last_mile")),
            "form_factor": _str(r.get("form_factor")),
            "max_weight": _float(r.get("max_weight")),
            "min_weight": _float(r.get("min_weight")),
            "max_length": _float(r.get("max_length")),
            "max_width": _float(r.get("max_width")),
            "max_height": _float(r.get("max_height")),
            "dimension_unit": _str(r.get("dimension_unit")),
            "weight_unit": _str(r.get("weight_unit")),
            "zone_ids": [],
            "surcharge_ids": _split_ids(r.get("surcharge_ids")),
        }
        for r in parsed.get("services", [])
        if _str(r.get("service_id"))
    ]

    svc_zone_map: dict = {}
    for sr in parsed.get("service_rates", []):
        sid = _str(sr.get("service_id"))
        zid = _str(sr.get("zone_id"))
        if sid and zid:
            svc_zone_map.setdefault(sid, set()).add(zid)

    for svc in services:
        svc_id = svc["id"]
        if svc_id in svc_zone_map:
            svc["zone_ids"] = list(svc_zone_map[svc_id])

    surcharges = [
        {
            "id": _str(r.get("surcharge_id")),
            "name": _str(r.get("surcharge_name")),
            "amount": _float(r.get("amount")) or 0.0,
            "surcharge_type": _str(r.get("surcharge_type")) or "fixed",
            "cost": _float(r.get("cost")),
            "active": _bool(r.get("active")) if r.get("active") is not None else True,
        }
        for r in parsed.get("surcharges", [])
        if _str(r.get("surcharge_id"))
    ]

    service_rates = [
        {
            "service_id": _str(r.get("service_id")),
            "zone_id": _str(r.get("zone_id")),
            "shipment_type": _str(r.get("shipment_type")) or "outbound",
            "min_weight": _float(r.get("min_weight")) or 0.0,
            "max_weight": _float(r.get("max_weight")) or 0.0,
            "rate": _float(r.get("base_rate")) or 0.0,
            "cost": _float(r.get("cost")),
            "transit_days": _str(r.get("transit_days")),
            "plan_rate_start": _float(r.get("plan_rate_start")),
            "plan_cost_start": _float(r.get("plan_cost_start")),
            "plan_rate_advanced": _float(r.get("plan_rate_advanced")),
            "plan_cost_advanced": _float(r.get("plan_cost_advanced")),
            "plan_rate_pro": _float(r.get("plan_rate_pro")),
            "plan_cost_pro": _float(r.get("plan_cost_pro")),
            "plan_rate_enterprise": _float(r.get("plan_rate_enterprise")),
            "plan_cost_enterprise": _float(r.get("plan_cost_enterprise")),
            "surcharge_ids": _split_ids(r.get("surcharge_ids")),
        }
        for r in parsed.get("service_rates", [])
        if _str(r.get("service_id")) and _str(r.get("zone_id"))
    ]

    return {
        "rate_sheet": {
            "name": _str(rs_row.get("name")),
            "carrier_name": _str(rs_row.get("carrier_name")),
            "slug": _str(rs_row.get("slug")),
            "currency": _str(rs_row.get("currency")),
            "origin_countries": _parse_countries(rs_row.get("origin_countries")),
        },
        "zones": zones,
        "services": services,
        "surcharges": surcharges,
        "service_rates": service_rates,
    }


# ─────────────────────────────────────────────────────────────────────────────
# DIFF COMPUTATION (for dry_run)
# ─────────────────────────────────────────────────────────────────────────────


def compute_diff(existing_sheet, payload: dict) -> dict:
    """
    Compare an existing RateSheet model instance against the incoming payload.

    Returns a diff dict:
    {
        "summary": {"updated": int, "added": int, "removed": int, "unchanged": int},
        "rows": [
            {
                "service_code": str,  "zone_label": str,
                "service_id": str,    "zone_id": str,   # aliases for backward-compat
                "shipment_type": str,
                "min_weight": float,  "max_weight": float,
                "old_rate": float|None,  "new_rate": float|None,
                "change": "updated"|"added"|"removed"|"unchanged"
            },
            ...
        ]
    }
    """
    if existing_sheet is None:
        rows = [
            {
                **_diff_row(
                    (
                        sr["service_id"],
                        sr["zone_id"],
                        sr.get("shipment_type", "outbound"),
                        float(sr.get("min_weight", 0) or 0),
                        float(sr.get("max_weight", 0) or 0),
                    )
                ),
                "old_rate": None,
                "new_rate": sr["rate"],
                "change": "added",
            }
            for sr in payload["service_rates"]
        ]
        return {
            "summary": {"updated": 0, "added": len(rows), "removed": 0, "unchanged": 0},
            "rows": rows,
        }

    def _rate_key(sr):
        return (
            sr.get("service_id"),
            sr.get("zone_id"),
            sr.get("shipment_type", "outbound"),
            float(sr.get("min_weight", 0) or 0),
            float(sr.get("max_weight", 0) or 0),
        )

    existing_map = {_rate_key(sr): sr for sr in (existing_sheet.service_rates or [])}
    incoming_map = {_rate_key(sr): sr for sr in payload["service_rates"]}

    rows = []
    added = updated = removed = unchanged = 0

    for key, new_sr in incoming_map.items():
        old_sr = existing_map.get(key)
        new_rate = new_sr.get("rate", 0)
        if old_sr is None:
            rows.append({**_diff_row(key), "old_rate": None, "new_rate": new_rate, "change": "added"})
            added += 1
        else:
            old_rate = float(old_sr.get("rate", 0) or 0)
            if abs(old_rate - float(new_rate or 0)) > 0.0001:
                rows.append({**_diff_row(key), "old_rate": old_rate, "new_rate": new_rate, "change": "updated"})
                updated += 1
            else:
                rows.append({**_diff_row(key), "old_rate": old_rate, "new_rate": new_rate, "change": "unchanged"})
                unchanged += 1

    for key, old_sr in existing_map.items():
        if key not in incoming_map:
            rows.append(
                {
                    **_diff_row(key),
                    "old_rate": float(old_sr.get("rate", 0) or 0),
                    "new_rate": None,
                    "change": "removed",
                }
            )
            removed += 1

    return {
        "summary": {"updated": updated, "added": added, "removed": removed, "unchanged": unchanged},
        "rows": rows,
    }


def _diff_row(key: tuple) -> dict:
    service_id, zone_id, shipment_type, min_weight, max_weight = key
    return {
        # New flat-format names
        "service_code": service_id,
        "zone_label": zone_id,
        # Legacy aliases (kept for backward-compat)
        "service_id": service_id,
        "zone_id": zone_id,
        "shipment_type": shipment_type,
        "min_weight": min_weight,
        "max_weight": max_weight,
    }


# ─────────────────────────────────────────────────────────────────────────────
# EXPORT — flat single-sheet format
# ─────────────────────────────────────────────────────────────────────────────

# Ordered column definitions: (column_name, column_width)
_FLAT_EXPORT_COLS = [
    ("carrier_name", 22),
    ("service_code", 22),
    ("carrier_service_code", 24),
    ("service_name", 30),
    ("shipment_type", 18),
    ("origin_country", 18),
    ("zone_label", 22),
    ("country_codes", 50),
    ("min_weight", 14),
    ("max_weight", 14),
    ("weight_unit", 14),
    ("max_length", 14),
    ("max_width", 13),
    ("max_height", 13),
    ("dimension_unit", 16),
    ("currency", 12),
    ("base_rate", 14),
    ("cost", 14),
    ("transit_days", 16),
    ("transit_time", 16),
    ("plan_rate_start", 20),
    ("plan_cost_start", 20),
    ("plan_rate_advanced", 22),
    ("plan_cost_advanced", 22),
    ("plan_rate_pro", 18),
    ("plan_cost_pro", 18),
    ("plan_rate_enterprise", 24),
    ("plan_cost_enterprise", 24),
    ("tracked", 12),
    ("b2c", 10),
    ("b2b", 10),
    ("first_mile", 18),
    ("last_mile", 18),
    ("form_factor", 18),
    ("signature", 12),
    ("fuel_surcharge", 18),
    ("seasonal_surcharge", 20),
    ("customs_surcharge", 20),
    ("energy_surcharge", 18),
    ("road_toll", 14),
    ("security_surcharge", 20),
    ("notes", 40),
]


def export_rate_sheet_xlsx(rate_sheet) -> bytes:
    """
    Export a RateSheet model instance to a flat single-sheet xlsx workbook.

    The workbook has one sheet (``service_rates``) with one fully-denormalised
    row per (carrier × service × zone × weight_range) combination.
    Returns raw bytes.
    """
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    NAVY = "1F3864"
    WHITE = "FFFFFF"
    LIGHT_GREY = "F2F2F2"

    hdr_font = Font(name="Calibri", bold=True, color=WHITE, size=10)
    hdr_fill = PatternFill("solid", fgColor=NAVY)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_font = Font(name="Calibri", size=10)
    cell_align = Alignment(horizontal="left", vertical="center")
    alt_fill = PatternFill("solid", fgColor=LIGHT_GREY)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "service_rates"
    ws.sheet_properties.tabColor = "1F3864"

    # Write header row
    col_names = [c for c, _ in _FLAT_EXPORT_COLS]
    for c, (name, width) in enumerate(_FLAT_EXPORT_COLS, 1):
        cell = ws.cell(row=1, column=c, value=name)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        ws.column_dimensions[get_column_letter(c)].width = width
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = ws.cell(row=2, column=1)

    # Build lookup tables from model
    services_by_id: dict = {}
    for svc in rate_sheet.services.all():
        svc_id = getattr(svc, "id", None) or getattr(svc, "service_code", None)
        if svc_id:
            services_by_id[svc_id] = svc

    zones_by_id: dict = {}
    for z in rate_sheet.zones or []:
        zid = z.get("id") or z.get("label")
        if zid:
            zones_by_id[zid] = z

    origin_country = ",".join(rate_sheet.origin_countries or [])

    data_row = 2
    for sr in rate_sheet.service_rates or []:
        svc = services_by_id.get(sr.get("service_id"))
        zone = zones_by_id.get(sr.get("zone_id"))

        features = {}
        if svc and hasattr(svc, "features") and isinstance(svc.features, dict):
            features = svc.features

        row_vals = {
            "carrier_name": rate_sheet.carrier_name,
            "service_code": svc.service_code if svc else sr.get("service_id"),
            "carrier_service_code": svc.carrier_service_code if svc else None,
            "service_name": svc.service_name if svc else None,
            "shipment_type": sr.get("shipment_type", "outbound"),
            "origin_country": origin_country,
            "zone_label": zone.get("label") if zone else sr.get("zone_id"),
            "country_codes": ",".join(zone.get("country_codes") or []) if zone else "",
            "min_weight": sr.get("min_weight"),
            "max_weight": sr.get("max_weight"),
            "weight_unit": (svc.weight_unit if svc and hasattr(svc, "weight_unit") else None) or "KG",
            "max_length": svc.max_length if svc and hasattr(svc, "max_length") else None,
            "max_width": svc.max_width if svc and hasattr(svc, "max_width") else None,
            "max_height": svc.max_height if svc and hasattr(svc, "max_height") else None,
            "dimension_unit": (svc.dimension_unit if svc and hasattr(svc, "dimension_unit") else None) or "CM",
            "currency": (svc.currency if svc and hasattr(svc, "currency") else None)
            or getattr(rate_sheet, "currency", "EUR")
            or "EUR",
            "base_rate": sr.get("rate"),
            "cost": sr.get("cost"),
            "transit_days": sr.get("transit_days") or (zone.get("transit_days") if zone else None),
            "transit_time": sr.get("transit_time"),
            "plan_rate_start": sr.get("plan_rate_start"),
            "plan_cost_start": sr.get("plan_cost_start"),
            "plan_rate_advanced": sr.get("plan_rate_advanced"),
            "plan_cost_advanced": sr.get("plan_cost_advanced"),
            "plan_rate_pro": sr.get("plan_rate_pro"),
            "plan_cost_pro": sr.get("plan_cost_pro"),
            "plan_rate_enterprise": sr.get("plan_rate_enterprise"),
            "plan_cost_enterprise": sr.get("plan_cost_enterprise"),
            "tracked": svc.tracked if svc and hasattr(svc, "tracked") else features.get("tracked"),
            "b2c": features.get("b2c"),
            "b2b": features.get("b2b"),
            "first_mile": features.get("first_mile"),
            "last_mile": features.get("last_mile"),
            "form_factor": features.get("form_factor"),
            "signature": None,
            "fuel_surcharge": None,
            "seasonal_surcharge": None,
            "customs_surcharge": None,
            "energy_surcharge": None,
            "road_toll": None,
            "security_surcharge": None,
            "notes": None,
        }

        fill = alt_fill if data_row % 2 == 0 else None
        for c, name in enumerate(col_names, 1):
            cell = ws.cell(row=data_row, column=c, value=row_vals.get(name))
            cell.font = cell_font
            cell.alignment = cell_align
            if fill:
                cell.fill = fill

        data_row += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# UPSERT LOGIC
# ─────────────────────────────────────────────────────────────────────────────


def upsert_rate_sheet(payload: dict, context) -> tuple:
    """
    Upsert a rate sheet from the parsed payload dict.

    Matches on slug → update if found, create if not.
    Returns (rate_sheet_instance, created: bool).
    """
    from karrio.server.providers.models import RateSheet, ServiceLevel

    rs_data = payload["rate_sheet"]
    slug = rs_data.get("slug")

    existing = RateSheet.access_by(context).filter(slug=slug).first() if slug else None

    if existing:
        existing.name = rs_data.get("name") or existing.name
        existing.carrier_name = rs_data.get("carrier_name") or existing.carrier_name
        if rs_data.get("origin_countries"):
            existing.origin_countries = rs_data["origin_countries"]
        existing.save(update_fields=["name", "carrier_name", "origin_countries", "updated_at"])
        sheet = existing
        created = False
    else:
        sheet = RateSheet(
            name=rs_data.get("name", ""),
            carrier_name=rs_data.get("carrier_name", ""),
            slug=slug or lib.uuid("rsht_"),
            origin_countries=rs_data.get("origin_countries", []),
            zones=[],
            surcharges=[],
            service_rates=[],
        )
        sheet.created_by = context.user
        if hasattr(context, "org") and context.org:
            sheet.org = context.org
        sheet.save()
        created = True

    sheet.zones = payload["zones"]
    sheet.surcharges = payload["surcharges"]
    sheet.service_rates = payload["service_rates"]
    sheet.save(update_fields=["zones", "surcharges", "service_rates"])

    incoming_svc_ids = {s["id"] for s in payload["services"]}
    existing_svcs = {svc.id: svc for svc in sheet.services.all()}

    for svc_data in payload["services"]:
        svc_id = svc_data["id"]
        if svc_id in existing_svcs:
            svc = existing_svcs[svc_id]
            svc.service_name = svc_data.get("service_name") or svc.service_name
            svc.service_code = svc_data.get("service_code") or svc.service_code
            svc.carrier_service_code = svc_data.get("carrier_service_code")
            svc.currency = svc_data.get("currency")
            svc.domicile = svc_data.get("domicile")
            svc.international = svc_data.get("international")
            svc.max_weight = svc_data.get("max_weight")
            svc.min_weight = svc_data.get("min_weight")
            svc.max_length = svc_data.get("max_length")
            svc.max_width = svc_data.get("max_width")
            svc.max_height = svc_data.get("max_height")
            svc.dimension_unit = svc_data.get("dimension_unit")
            svc.weight_unit = svc_data.get("weight_unit")
            svc.zone_ids = svc_data.get("zone_ids", [])
            svc.surcharge_ids = svc_data.get("surcharge_ids", [])
            svc.save()
        else:
            new_svc = ServiceLevel(
                id=svc_id,
                service_name=svc_data.get("service_name", ""),
                service_code=svc_data.get("service_code", ""),
                carrier_service_code=svc_data.get("carrier_service_code"),
                currency=svc_data.get("currency"),
                domicile=svc_data.get("domicile"),
                international=svc_data.get("international"),
                max_weight=svc_data.get("max_weight"),
                min_weight=svc_data.get("min_weight"),
                max_length=svc_data.get("max_length"),
                max_width=svc_data.get("max_width"),
                max_height=svc_data.get("max_height"),
                dimension_unit=svc_data.get("dimension_unit"),
                weight_unit=svc_data.get("weight_unit"),
                zone_ids=svc_data.get("zone_ids", []),
                surcharge_ids=svc_data.get("surcharge_ids", []),
            )
            if hasattr(sheet, "created_by") and sheet.created_by:
                new_svc.created_by = sheet.created_by
            if hasattr(sheet, "org") and sheet.org:
                new_svc.org = sheet.org
            new_svc.save()
            sheet.services.add(new_svc)

    for svc_id, svc in existing_svcs.items():
        if svc_id not in incoming_svc_ids:
            sheet.services.remove(svc)

    return sheet, created
